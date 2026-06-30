import io
import json
from datetime import date
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session
from app.models.all_models import Employee, CareerHistory, DisciplinaryAction, Leave, User, FinancialRevenue, FinancialExpense, Supplier

def export_consolidated_data_excel(db: Session, current_user: User) -> io.BytesIO:
    """
    Exports a comprehensive multi-tab Excel spreadsheet containing:
    1. Employee master database
    2. Career evolution history
    3. Disciplinary actions history
    4. Leaves and absences history
    """
    wb = Workbook()
    
    # Setup styling tokens
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    data_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Slate
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Light Grey
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # ----------------- TAB 1: COLABORADORES -----------------
    ws1 = wb.active
    ws1.title = "Colaboradores"
    ws1.views.sheetView[0].showGridLines = True
    
    # Headers
    headers1 = [
        "Matrícula", "Nome Completo", "CPF", "RG", "CTPS", "PIS", "Reservista", "Data Nascimento", "E-mail", "Telefone",
        "Cargo", "Departamento", "Data Admissão", "Salário Base", "Status", "Data Demissão", "Justificativa"
    ]
    
    ws1.append(["MEURESTÔ - GESTÃO DE PESSOAL"])
    ws1.cell(1, 1).font = title_font
    ws1.append([]) # Empty spacer
    ws1.append(headers1)
    
    # Set header styles
    for col_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Data rows
    query1 = db.query(Employee)
    if current_user.role != "admin":
        query1 = query1.filter(Employee.group_id == current_user.group_id)
    employees = query1.all()
    
    for row_idx, emp in enumerate(employees, start=4):
        adm_date = emp.contract.admission_date.strftime("%d/%m/%Y") if emp.contract and emp.contract.admission_date else ""
        salary = emp.contract.base_salary if emp.contract else 0.0
        role = emp.contract.role if emp.contract else ""
        dept = emp.contract.department if emp.contract else ""
        dob = emp.dob.strftime("%d/%m/%Y") if emp.dob else ""
        term_date = emp.termination_date.strftime("%d/%m/%Y") if emp.termination_date else ""
        
        row_data = [
            emp.registration_number, emp.name, emp.cpf, emp.rg, emp.ctps or "", emp.pis or "", emp.reservista or "", dob, emp.email, emp.phone,
            role, dept, adm_date, salary, (emp.status.upper() if emp.status else 'ATIVO'), term_date, emp.termination_reason or ""
        ]
        ws1.append(row_data)
        
        # Apply data cell styling
        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = alt_row_fill
            # Alignments
            if col_idx in [1, 3, 4, 5, 6, 7, 8, 13, 15, 16]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 14:
                cell.number_format = '"R$" #,##0.00'
                cell.alignment = Alignment(horizontal="right")
                
    # Auto-fit columns
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
        
    # ----------------- TAB 2: HISTÓRICO FUNCIONAL -----------------
    ws2 = wb.create_sheet(title="Histórico Funcional")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = ["Matrícula Colab.", "Nome Colaborador", "Data Alteração", "Campo Alterado", "Valor Antigo", "Valor Novo", "Motivo"]
    ws2.append(["EVOLUÇÃO E ALTERAÇÕES CONTRATUAIS"])
    ws2.cell(1, 1).font = title_font
    ws2.append([])
    ws2.append(headers2)
    
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    query2 = db.query(CareerHistory).join(Employee)
    if current_user.role != "admin":
        query2 = query2.filter(Employee.group_id == current_user.group_id)
    histories = query2.order_by(CareerHistory.change_date.desc()).all()
    
    for row_idx, hist in enumerate(histories, start=4):
        emp_name = hist.employee.name if hist.employee else "N/A"
        emp_reg = hist.employee.registration_number if hist.employee else "N/A"
        chg_date = hist.change_date.strftime("%d/%m/%Y %H:%M")
        
        row_data = [emp_reg, emp_name, chg_date, (hist.field_name.upper() if hist.field_name else ''), hist.old_value, hist.new_value, hist.reason]
        ws2.append(row_data)
        
        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = alt_row_fill
            if col_idx in [1, 3, 4]:
                cell.alignment = Alignment(horizontal="center")
                
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
        
    # ----------------- TAB 3: AUDITORIA DISCIPLINAR -----------------
    ws3 = wb.create_sheet(title="Histórico Disciplinar")
    ws3.views.sheetView[0].showGridLines = True
    
    headers3 = ["Matrícula", "Colaborador", "Data Ocorrência", "Tipo de Ação", "Motivo/Infração", "Detalhes", "Gestor Responsável"]
    ws3.append(["REGISTROS E OCORRÊNCIAS DISCIPLINARES"])
    ws3.cell(1, 1).font = title_font
    ws3.append([])
    ws3.append(headers3)
    
    for col_idx in range(1, len(headers3) + 1):
        cell = ws3.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    query3 = db.query(DisciplinaryAction).join(Employee)
    if current_user.role != "admin":
        query3 = query3.filter(Employee.group_id == current_user.group_id)
    disc_actions = query3.order_by(DisciplinaryAction.action_date.desc()).all()
    
    for row_idx, act in enumerate(disc_actions, start=4):
        emp_name = act.employee.name if act.employee else "N/A"
        emp_reg = act.employee.registration_number if act.employee else "N/A"
        act_date = act.action_date.strftime("%d/%m/%Y")
        
        row_data = [emp_reg, emp_name, act_date, (act.type.upper() if act.type else ''), act.reason, act.details, act.manager_name]
        ws3.append(row_data)
        
        for col_idx in range(1, len(headers3) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = alt_row_fill
            if col_idx in [1, 3, 4]:
                cell.alignment = Alignment(horizontal="center")
                
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
        
    # ----------------- TAB 4: AFASTAMENTOS -----------------
    ws4 = wb.create_sheet(title="Afastamentos")
    ws4.views.sheetView[0].showGridLines = True
    
    headers4 = ["Matrícula", "Colaborador", "Data Início", "Data Fim", "Motivo / Descrição"]
    ws4.append(["CONTROLE DE LICENÇAS E AFASTAMENTOS"])
    ws4.cell(1, 1).font = title_font
    ws4.append([])
    ws4.append(headers4)
    
    for col_idx in range(1, len(headers4) + 1):
        cell = ws4.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    query4 = db.query(Leave).join(Employee)
    if current_user.role != "admin":
        query4 = query4.filter(Employee.group_id == current_user.group_id)
    leaves = query4.order_by(Leave.start_date.desc()).all()
    
    for row_idx, leave in enumerate(leaves, start=4):
        emp_name = leave.employee.name if leave.employee else "N/A"
        emp_reg = leave.employee.registration_number if leave.employee else "N/A"
        start = leave.start_date.strftime("%d/%m/%Y")
        end = leave.end_date.strftime("%d/%m/%Y")
        
        row_data = [emp_reg, emp_name, start, end, leave.reason]
        ws4.append(row_data)
        
        for col_idx in range(1, len(headers4) + 1):
            cell = ws4.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = alt_row_fill
            if col_idx in [1, 3, 4]:
                cell.alignment = Alignment(horizontal="center")
                
    for col in ws4.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws4.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Save Workbook to Byte buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_financial_excel(
    db: Session,
    current_user: User,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    category: Optional[str] = None,
    supplier_id: Optional[str] = None,
    payment_method: Optional[str] = None,
    status: Optional[str] = None
) -> io.BytesIO:
    """
    Exports filtered financial data into a multi-tab spreadsheet.
    """
    wb = Workbook()
    
    # Setup styling tokens
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    data_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Fetch Revenues (Recebimentos)
    rev_q = db.query(FinancialRevenue)
    if current_user.role != "admin":
        rev_q = rev_q.filter(FinancialRevenue.group_id == current_user.group_id)
        
    if start_date:
        rev_q = rev_q.filter(FinancialRevenue.date >= start_date)
    if end_date:
        rev_q = rev_q.filter(FinancialRevenue.date <= end_date)
    if category:
        rev_q = rev_q.filter(FinancialRevenue.category == category)
    if payment_method:
        rev_q = rev_q.filter(FinancialRevenue.payment_method == payment_method)
    if status:
        rev_q = rev_q.filter(FinancialRevenue.status == status)
        
    revenues = rev_q.order_by(FinancialRevenue.date.desc()).all()
    
    # Fetch Expenses (Pagamentos)
    exp_q = db.query(FinancialExpense)
    if current_user.role != "admin":
        exp_q = exp_q.filter(FinancialExpense.group_id == current_user.group_id)
        
    if start_date:
        exp_q = exp_q.filter(FinancialExpense.date >= start_date)
    if end_date:
        exp_q = exp_q.filter(FinancialExpense.date <= end_date)
    if category:
        exp_q = exp_q.filter(FinancialExpense.category == category)
    if supplier_id:
        exp_q = exp_q.filter(FinancialExpense.supplier_id == supplier_id)
    if payment_method:
        exp_q = exp_q.filter(FinancialExpense.payment_method == payment_method)
    if status:
        exp_q = exp_q.filter(FinancialExpense.status == status)
        
    expenses = exp_q.order_by(FinancialExpense.date.desc()).all()

    # ----------------- TAB 1: FLUXO FINANCEIRO -----------------
    ws1 = wb.active
    ws1.title = "Fluxo Financeiro"
    ws1.views.sheetView[0].showGridLines = True
    
    headers1 = ["Data", "Tipo", "Descrição", "Cliente/Fornecedor", "Categoria", "Valor (R$)", "Forma de Pagamento", "Status"]
    
    ws1.append(["MEURESTÔ - FLUXO FINANCEIRO CONSOLIDADO"])
    ws1.cell(1, 1).font = title_font
    ws1.append([])
    ws1.append(headers1)
    
    # Style headers
    for col_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    row_idx = 4
    # Merge both into a single sorted chronological flow
    flow_items = []
    for r in revenues:
        flow_items.append((r.date, "RECEITA", r.description, r.client or "N/A", r.category, r.amount, r.payment_method or "N/A", r.status))
    for e in expenses:
        supp_name = e.supplier.trade_name if e.supplier else "N/A"
        flow_items.append((e.date, "DESPESA", e.description, supp_name, e.category, e.amount, e.payment_method or "N/A", e.status))
        
    flow_items.sort(key=lambda x: x[0], reverse=True)
    
    for item in flow_items:
        date_str = item[0].strftime("%d/%m/%Y") if item[0] else ""
        row_data = [date_str, item[1], item[2], item[3], item[4], item[5], item[6], (item[7].upper() if item[7] else 'N/A')]
        ws1.append(row_data)
        
        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = alt_row_fill
                
            if col_idx in [1, 2, 7, 8]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 6:
                cell.number_format = '"R$" #,##0.00'
                cell.alignment = Alignment(horizontal="right")
                
        row_idx += 1
        
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ----------------- TAB 2: CONTAS A RECEBER -----------------
    ws2 = wb.create_sheet(title="Contas a Receber")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = ["Data Prevista", "Descrição", "Cliente", "Categoria", "Valor (R$)", "Forma de Recebimento", "Status"]
    ws2.append(["MEURESTÔ - CONTAS A RECEBER (PREVISÕES E PENDÊNCIAS)"])
    ws2.cell(1, 1).font = title_font
    ws2.append([])
    ws2.append(headers2)
    
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    row_idx = 4
    for r in revenues:
        if r.status == "Recebido":
            continue
        date_str = r.expected_date.strftime("%d/%m/%Y") if r.expected_date else (r.date.strftime("%d/%m/%Y") if r.date else "")
        row_data = [date_str, r.description, r.client or "N/A", r.category, r.amount, r.payment_method or "N/A", (r.status.upper() if r.status else 'N/A')]
        ws2.append(row_data)
        
        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = alt_row_fill
            if col_idx in [1, 6, 7]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 5:
                cell.number_format = '"R$" #,##0.00'
                cell.alignment = Alignment(horizontal="right")
        row_idx += 1
        
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ----------------- TAB 3: CONTAS A PAGAR -----------------
    ws3 = wb.create_sheet(title="Contas a Pagar")
    ws3.views.sheetView[0].showGridLines = True
    
    headers3 = ["Vencimento", "Descrição", "Fornecedor", "Categoria", "Valor (R$)", "Forma de Pagamento", "Status", "Recorrente?"]
    ws3.append(["MEURESTÔ - CONTAS A PAGAR (PENDÊNCIAS E OBRIGAÇÕES)"])
    ws3.cell(1, 1).font = title_font
    ws3.append([])
    ws3.append(headers3)
    
    for col_idx in range(1, len(headers3) + 1):
        cell = ws3.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    row_idx = 4
    for e in expenses:
        if e.status == "Pago":
            continue
        supp_name = e.supplier.trade_name if e.supplier else "N/A"
        date_str = e.due_date.strftime("%d/%m/%Y") if e.due_date else (e.date.strftime("%d/%m/%Y") if e.date else "")
        recur_str = f"Sim ({e.recurrence_period})" if e.is_recurring else "Não"
        row_data = [date_str, e.description, supp_name, e.category, e.amount, e.payment_method or "N/A", (e.status.upper() if e.status else 'N/A'), recur_str]
        ws3.append(row_data)
        
        for col_idx in range(1, len(headers3) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = alt_row_fill
            if col_idx in [1, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 5:
                cell.number_format = '"R$" #,##0.00'
                cell.alignment = Alignment(horizontal="right")
        row_idx += 1
        
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_employee_dossier_excel(db: Session, employee_id: str, current_user: User) -> io.BytesIO:
    """
    Exports a comprehensive multi-tab Excel spreadsheet for a single employee containing:
    1. Cadastro Geral (Basic Data, Address, Mother/Father, Dependents, Bank details)
    2. Contrato & Jornada (Contract info and Shifts)
    3. Histórico de Alterações (Career updates)
    4. Licenças e Atestados (Leaves)
    5. Ocorrências Disciplinares (Disciplinary actions)
    """
    # Fetch employee
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise ValueError("Colaborador não encontrado.")

    # Enforce tenant isolation
    if current_user.role != "admin" and emp.group_id != current_user.group_id:
        raise PermissionError("Acesso não autorizado aos dados desta empresa.")

    wb = Workbook()
    
    # Styling tokens
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    section_font = Font(name="Calibri", size=12, bold=True, color="d97706")
    data_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # ----------------- TAB 1: CADASTRO GERAL -----------------
    ws1 = wb.active
    ws1.title = "Cadastro Geral"
    ws1.views.sheetView[0].showGridLines = True

    ws1.append([f"DOSSIÊ DO COLABORADOR - {(emp.name.upper() if emp.name else '')}"])
    ws1.cell(1, 1).font = title_font
    ws1.append([])

    # Basic Info
    ws1.append(["DADOS PESSOAIS"])
    ws1.cell(3, 1).font = section_font
    
    headers_basic = ["Matrícula", "CPF", "RG", "Sexo", "Data Nasc.", "Estado Civil", "Nacionalidade", "E-mail", "Telefone"]
    ws1.append(headers_basic)
    for col in range(1, len(headers_basic) + 1):
        cell = ws1.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    dob = emp.dob.strftime("%d/%m/%Y") if emp.dob else ""
    row_basic = [
        emp.registration_number, emp.cpf, emp.rg, emp.sex or "N/A", dob, 
        emp.civil_status, emp.nationality, emp.email, emp.phone
    ]
    ws1.append(row_basic)
    for col in range(1, len(headers_basic) + 1):
        cell = ws1.cell(row=5, column=col)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    ws1.append([]) # Spacer

    # Address & Parents
    ws1.append(["ENDEREÇO E FILIAÇÃO"])
    ws1.cell(7, 1).font = section_font

    headers_addr = ["CEP", "Logradouro", "Número", "Complemento", "Bairro", "Cidade", "UF", "Nome da Mãe", "Nome do Pai"]
    ws1.append(headers_addr)
    for col in range(1, len(headers_addr) + 1):
        cell = ws1.cell(row=8, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_addr = [
        emp.address_cep, emp.address_street, emp.address_number, emp.address_complement or "", 
        emp.address_neighborhood, emp.address_city, emp.address_state, emp.mother_name, emp.father_name or ""
    ]
    ws1.append(row_addr)
    for col in range(1, len(headers_addr) + 1):
        cell = ws1.cell(row=9, column=col)
        cell.font = data_font
        cell.border = thin_border
        if col in [1, 3, 7]:
            cell.alignment = Alignment(horizontal="center")

    ws1.append([]) # Spacer

    # Bank info & Docs info
    ws1.append(["DADOS BANCÁRIOS E DOCUMENTAÇÃO COMPLEMENTAR"])
    ws1.cell(11, 1).font = section_font
    
    headers_bank = ["Banco", "Agência", "Conta", "Chave Pix", "CTPS", "PIS", "Certificado Reservista"]
    ws1.append(headers_bank)
    for col in range(1, len(headers_bank) + 1):
        cell = ws1.cell(row=12, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_bank = [
        emp.bank_name or "N/A", emp.bank_agency or "N/A", emp.bank_account or "N/A", emp.pix_key or "N/A",
        emp.ctps or "N/A", emp.pis or "N/A", emp.reservista or "N/A"
    ]
    ws1.append(row_bank)
    for col in range(1, len(headers_bank) + 1):
        cell = ws1.cell(row=13, column=col)
        cell.font = data_font
        cell.border = thin_border
        if col in [2, 3, 5, 6, 7]:
            cell.alignment = Alignment(horizontal="center")

    ws1.append([]) # Spacer

    # Dependents
    ws1.append(["DEPENDENTES"])
    ws1.cell(15, 1).font = section_font

    headers_dep = ["Nome do Dependente", "Grau de Parentesco", "Data de Nascimento"]
    ws1.append(headers_dep)
    for col in range(1, len(headers_dep) + 1):
        cell = ws1.cell(row=16, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    if not emp.dependents:
        ws1.append(["Sem dependentes cadastrados", "", ""])
        for col in range(1, len(headers_dep) + 1):
            cell = ws1.cell(row=17, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    else:
        for idx, dep in enumerate(emp.dependents, start=17):
            dep_dob = dep.dob.strftime("%d/%m/%Y") if dep.dob else ""
            ws1.append([dep.name, dep.relationship, dep_dob])
            for col in range(1, len(headers_dep) + 1):
                cell = ws1.cell(row=idx, column=col)
                cell.font = data_font
                cell.border = thin_border
                if col in [2, 3]:
                    cell.alignment = Alignment(horizontal="center")

    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ----------------- TAB 2: CONTRATO E JORNADA -----------------
    ws2 = wb.create_sheet(title="Contrato e Jornada")
    ws2.views.sheetView[0].showGridLines = True

    ws2.append(["INFORMAÇÕES CONTRATUAIS E DE JORNADA"])
    ws2.cell(1, 1).font = title_font
    ws2.append([])

    # Contract Details
    ws2.append(["DADOS DO CONTRATO"])
    ws2.cell(3, 1).font = section_font

    headers_contract = ["Cargo", "Departamento/Setor", "Data Admissão", "Salário Base", "Status do Contrato", "Tipo Contrato"]
    ws2.append(headers_contract)
    for col in range(1, len(headers_contract) + 1):
        cell = ws2.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    if emp.contract:
        adm_date = emp.contract.admission_date.strftime("%d/%m/%Y") if emp.contract.admission_date else ""
        row_contract = [
            emp.contract.role, emp.contract.department, adm_date, 
            emp.contract.base_salary, (emp.status.upper() if emp.status else 'ATIVO'), emp.contract.contract_type or "N/A"
        ]
    else:
        row_contract = ["N/A", "N/A", "N/A", 0.0, (emp.status.upper() if emp.status else 'ATIVO'), "N/A"]

    ws2.append(row_contract)
    for col in range(1, len(headers_contract) + 1):
        cell = ws2.cell(row=5, column=col)
        cell.font = data_font
        cell.border = thin_border
        if col in [3, 5, 6]:
            cell.alignment = Alignment(horizontal="center")
        elif col == 4:
            cell.number_format = '"R$" #,##0.00'
            cell.alignment = Alignment(horizontal="right")

    ws2.append([]) # Spacer

    # Shift / Jornada Details
    ws2.append(["JORNADA DE TRABALHO"])
    ws2.cell(7, 1).font = section_font

    headers_shift = ["Escala", "Horário de Entrada", "Horário de Saída", "Intervalo (Minutos)", "Banco de Horas (Minutos)"]
    ws2.append(headers_shift)
    for col in range(1, len(headers_shift) + 1):
        cell = ws2.cell(row=8, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    if emp.shift:
        row_shift = [
            emp.shift.scale_type, emp.shift.entry_time, emp.shift.exit_time, 
            emp.shift.interval_duration_minutes, emp.shift.bank_of_hours_minutes
        ]
    else:
        row_shift = ["Não cadastrada", "N/A", "N/A", 0, 0]

    ws2.append(row_shift)
    for col in range(1, len(headers_shift) + 1):
        cell = ws2.cell(row=9, column=col)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    if emp.status == "terminated":
        ws2.append([])
        ws2.append(["DADOS DE DESLIGAMENTO"])
        ws2.cell(ws2.max_row, 1).font = section_font
        
        headers_term = ["Data de Demissão", "Justificativa / Motivo"]
        ws2.append(headers_term)
        header_row = ws2.max_row
        for col in range(1, len(headers_term) + 1):
            cell = ws2.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        term_date_str = emp.termination_date.strftime("%d/%m/%Y") if emp.termination_date else "Não informada"
        ws2.append([term_date_str, emp.termination_reason or "N/A"])
        data_row = ws2.max_row
        for col in range(1, len(headers_term) + 1):
            cell = ws2.cell(row=data_row, column=col)
            cell.font = data_font
            cell.border = thin_border
            if col == 1:
                cell.alignment = Alignment(horizontal="center")

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ----------------- TAB 3: HISTÓRICO DE ALTERAÇÕES -----------------
    ws3 = wb.create_sheet(title="Histórico de Alterações")
    ws3.views.sheetView[0].showGridLines = True

    ws3.append(["HISTÓRICO DE EVOLUÇÃO E ALTERAÇÕES CONTRATUAIS"])
    ws3.cell(1, 1).font = title_font
    ws3.append([])

    headers_history = ["Data da Alteração", "Campo Alterado", "Valor Antigo", "Valor Novo", "Motivo", "Responsável"]
    ws3.append(headers_history)
    for col in range(1, len(headers_history) + 1):
        cell = ws3.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    if not emp.career_history:
        ws3.append(["Sem alterações registradas no histórico", "", "", "", "", ""])
        for col in range(1, len(headers_history) + 1):
            cell = ws3.cell(row=4, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    else:
        sorted_history = sorted(emp.career_history, key=lambda x: x.change_date, reverse=True)
        for idx, hist in enumerate(sorted_history, start=4):
            chg_date = hist.change_date.strftime("%d/%m/%Y %H:%M")
            ws3.append([chg_date, (hist.field_name.upper() if hist.field_name else ''), hist.old_value or "", hist.new_value or "", hist.reason or "", hist.changed_by_username])
            for col in range(1, len(headers_history) + 1):
                cell = ws3.cell(row=idx, column=col)
                cell.font = data_font
                cell.border = thin_border
                if col in [1, 2, 6]:
                    cell.alignment = Alignment(horizontal="center")
                if idx % 2 == 1:
                    cell.fill = alt_row_fill

    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ----------------- TAB 4: ATESTADOS E AFASTAMENTOS -----------------
    ws4 = wb.create_sheet(title="Afastamentos e Atestados")
    ws4.views.sheetView[0].showGridLines = True

    ws4.append(["REGISTRO DE LICENÇAS, ATESTADOS E AFASTAMENTOS"])
    ws4.cell(1, 1).font = title_font
    ws4.append([])

    headers_leaves = ["Data de Início", "Data de Fim", "Motivo / Diagnóstico", "Comprovante Anexado"]
    ws4.append(headers_leaves)
    for col in range(1, len(headers_leaves) + 1):
        cell = ws4.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    if not emp.leaves:
        ws4.append(["Sem registros de afastamentos ou atestados", "", "", ""])
        for col in range(1, len(headers_leaves) + 1):
            cell = ws4.cell(row=4, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    else:
        sorted_leaves = sorted(emp.leaves, key=lambda x: x.start_date, reverse=True)
        for idx, leave in enumerate(sorted_leaves, start=4):
            start = leave.start_date.strftime("%d/%m/%Y") if leave.start_date else ""
            end = leave.end_date.strftime("%d/%m/%Y") if leave.end_date else ""
            has_cert = "Sim" if leave.doctor_certificate_path else "Não"
            ws4.append([start, end, leave.reason, has_cert])
            for col in range(1, len(headers_leaves) + 1):
                cell = ws4.cell(row=idx, column=col)
                cell.font = data_font
                cell.border = thin_border
                if col in [1, 2, 4]:
                    cell.alignment = Alignment(horizontal="center")
                if idx % 2 == 1:
                    cell.fill = alt_row_fill

    for col in ws4.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws4.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ----------------- TAB 5: OCORRÊNCIAS DISCIPLINARES -----------------
    ws5 = wb.create_sheet(title="Ocorrências Disciplinares")
    ws5.views.sheetView[0].showGridLines = True

    ws5.append(["REGISTRO DE OCORRÊNCIAS E MEDIDAS DISCIPLINARES"])
    ws5.cell(1, 1).font = title_font
    ws5.append([])

    headers_disc = ["Data da Ocorrência", "Tipo de Ação", "Motivo / Infração", "Detalhes da Ação", "Duração (Dias)", "Gestor Responsável"]
    ws5.append(headers_disc)
    for col in range(1, len(headers_disc) + 1):
        cell = ws5.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    if not emp.disciplinary_actions:
        ws5.append(["Sem ocorrências disciplinares registradas", "", "", "", "", ""])
        for col in range(1, len(headers_disc) + 1):
            cell = ws5.cell(row=4, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    else:
        sorted_disc = sorted(emp.disciplinary_actions, key=lambda x: x.action_date, reverse=True)
        for idx, action in enumerate(sorted_disc, start=4):
            act_date = action.action_date.strftime("%d/%m/%Y") if action.action_date else ""
            dur = action.duration_days if action.duration_days else "-"
            
            type_lbl = (action.type.upper() if action.type else '')
            if action.type == "warning": type_lbl = "ADVERTÊNCIA"
            elif action.type == "suspension": type_lbl = "SUSPENSÃO"
            elif action.type == "termination": type_lbl = "DESLIGAMENTO"
            
            ws5.append([act_date, type_lbl, action.reason, action.details, dur, action.manager_name])
            for col in range(1, len(headers_disc) + 1):
                cell = ws5.cell(row=idx, column=col)
                cell.font = data_font
                cell.border = thin_border
                if col in [1, 2, 5, 6]:
                    cell.alignment = Alignment(horizontal="center")
                if idx % 2 == 1:
                    cell.fill = alt_row_fill

    for col in ws5.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws5.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


